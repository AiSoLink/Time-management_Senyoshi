"""
対面アルコール ＋ 遠隔アルコール → 乗務員ID順・日時昇順の統合イベント列。
遠隔と対面の紐づけは行わず、1イベント＝(乗務員ID, 乗務員名, 日時, 種別)。
種別は「出庫」「帰庫」。デジタコとの突合は出庫同士・帰庫同士で ±120分。
"""
from __future__ import annotations

import csv
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


# 対面: 列名で取得 A=ID, B=氏名, E=日時, H=出帰庫
# 遠隔: 列番号 A=社員コード, B=社員名, E=出庫日時, Z=帰庫日時
ENKAKU_COL_INDEX = {"staff_code": 0, "staff_name": 1, "departure": 4, "return_": 25}

# デジタコ日時フォーマット（pipeline と揃える）
DT_FMT = "%Y/%m/%d %H:%M"


def _cell_value(ws: Worksheet, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


# 2026/08/25 のアルキラーNEXアップデートで「日時」→「検知日時」に変更されるため両方受け付ける
DT_HEADER_NAMES = ("日時", "検知日時")


def _find_header_row(ws: Worksheet, max_rows: int = 50) -> Optional[int]:
    want = {"ID", "氏名", "出帰庫"}
    for r in range(1, min(max_rows, ws.max_row + 1) + 1):
        row_vals = [_cell_value(ws, r, c) for c in range(1, 30)]
        cells_norm = {_normalize_header(str(v) if v is not None else "") for v in (row_vals[:20] if row_vals else [])}
        if want <= cells_norm and any(h in cells_norm for h in DT_HEADER_NAMES):
            return r
    return None


def _col_index_by_name(ws: Worksheet, header_row: int, name: str) -> Optional[int]:
    name_norm = _normalize_header(name)
    for c in range(1, ws.max_column + 1):
        if _normalize_header(str(_cell_value(ws, header_row, c) or "")) == name_norm:
            return c
    return None


def _normalize_header(s: Optional[str]) -> str:
    """ヘッダー比較用: 前後空白・ノーブレークスペース・NFKC正規化。' ID' や 'ID ' を 'ID' に揃える。"""
    if not s:
        return ""
    s = str(s).replace("\u00a0", " ").replace("\u3000", " ")  # NBSP・全角スペース→半角
    s = unicodedata.normalize("NFKC", s).strip()
    return s


def _normalize_dep_return(typ: Any) -> Optional[str]:
    """対面の「出帰庫」列を正規化。出庫/帰庫の表記ゆれ（全角・前後空白・接尾辞）を許容し '出庫' or '帰庫' を返す。"""
    if typ is None:
        return None
    s = unicodedata.normalize("NFKC", str(typ).replace("\u00a0", " ").replace("\u3000", " ").strip())
    if not s:
        return None
    if s == "出庫" or (s.startswith("出庫") and len(s) >= 2):
        return "出庫"
    if s == "帰庫" or (s.startswith("帰庫") and len(s) >= 2):
        return "帰庫"
    return None


def _read_csv_rows(path: Path, max_rows: int = 10000) -> List[List[str]]:
    """CSVを読み、行リストで返す。日本語Excel由来は cp932 が多いので先に試す。"""
    for enc in ("cp932", "utf-8-sig", "utf-8"):
        try:
            text = path.read_text(encoding=enc)
            rows: List[List[str]] = []
            for row in csv.reader(text.splitlines(), dialect=csv.excel):
                rows.append([_normalize_header(c) for c in row])
                if len(rows) >= max_rows:
                    break
            return rows
        except (UnicodeDecodeError, Exception):
            continue
    return []


def _to_datetime(v: Any) -> Optional[datetime]:
    """Excel日時 or 文字列を datetime に。解釈できない場合は None。"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    # 対面CSVの "2025/10/1 4:51" のように月・日・時が1桁の場合は2桁に揃える
    s = re.sub(r"/(\d)(?=/)", r"/0\1", s)   # 月: 2025/1/ -> 2025/01/
    s = re.sub(r"/(\d)(?=\s)", r"/0\1", s)  # 日: /1  -> /01
    s = re.sub(r" (\d)(?=:)", r" 0\1", s)   # 時:  4: ->  04:
    try:
        return datetime.strptime(s, DT_FMT)
    except ValueError:
        pass
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def load_taimen_sheet(ws: Worksheet) -> List[Tuple[Any, Any, Any, str]]:
    """対面シートから (ID, 氏名, 日時, 種別) のリスト。1行＝1イベント。"""
    header_row = _find_header_row(ws)
    if header_row is None:
        return []

    col_id = _col_index_by_name(ws, header_row, "ID")
    col_name = _col_index_by_name(ws, header_row, "氏名")
    col_dt = None
    for _h in DT_HEADER_NAMES:
        col_dt = _col_index_by_name(ws, header_row, _h)
        if col_dt:
            break
    col_type = _col_index_by_name(ws, header_row, "出帰庫")
    if not all([col_id, col_name, col_dt, col_type]):
        return []

    rows: List[Tuple[Any, Any, Any, str]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        id_ = _cell_value(ws, r, col_id)
        name = _cell_value(ws, r, col_name)
        dt = _cell_value(ws, r, col_dt)
        typ = _cell_value(ws, r, col_type)
        if id_ is None and name is None and dt is None:
            continue
        typ_str = _normalize_dep_return(typ)
        if typ_str is None:
            continue
        rows.append((id_, name, dt, typ_str))
    return rows


def _find_header_row_csv(rows: List[List[str]], col_names: List[str]) -> Optional[int]:
    """CSVの行リストから、指定列名が含まれるヘッダー行の0-basedインデックスを返す。正規化して比較。"""
    want = {_normalize_header(h) for h in col_names}
    for i, row in enumerate(rows[:50]):
        cells_norm = [_normalize_header(c) for c in row]
        if want <= set(cells_norm):  # 欲しい列名がすべてこの行に含まれる
            return i
    return None


def _col_index_by_name_csv(header_row: List[str], name: str) -> Optional[int]:
    """ヘッダー行から列名の0-basedインデックス。正規化して比較（' ID' も 'ID' と一致）。"""
    name_norm = _normalize_header(name)
    for j, c in enumerate(header_row):
        if _normalize_header(c) == name_norm:
            return j
    return None


def load_taimen_csv(path: Path) -> List[Tuple[Any, Any, Any, str]]:
    """対面CSVから (ID, 氏名, 日時, 種別) のリスト。列名で取得。日時列は「日時」「検知日時」両対応。"""
    rows = _read_csv_rows(path)
    hi = None
    for dt_name in DT_HEADER_NAMES:
        hi = _find_header_row_csv(rows, ["ID", "氏名", dt_name, "出帰庫"])
        if hi is not None:
            break
    if hi is None:
        return []
    header = rows[hi]
    col_id = _col_index_by_name_csv(header, "ID")
    col_name = _col_index_by_name_csv(header, "氏名")
    col_dt = None
    for dt_name in DT_HEADER_NAMES:
        col_dt = _col_index_by_name_csv(header, dt_name)
        if col_dt is not None:
            break
    col_type = _col_index_by_name_csv(header, "出帰庫")
    if col_id is None or col_name is None or col_dt is None or col_type is None:
        return []
    out: List[Tuple[Any, Any, Any, str]] = []
    for row in rows[hi + 1 :]:
        if len(row) <= max(col_id, col_name, col_dt, col_type):
            continue
        id_ = row[col_id] if col_id < len(row) else ""
        name = row[col_name] if col_name < len(row) else ""
        dt = row[col_dt] if col_dt < len(row) else ""
        typ = (row[col_type] or "").strip() if col_type < len(row) else ""
        if not id_ and not name and not dt:
            continue
        typ_str = _normalize_dep_return(typ)
        if typ_str is None:
            continue
        out.append((id_ or None, name or None, dt or None, typ_str))
    return out


def load_taimen_events(taimen_dir: Path) -> List[Tuple[Any, Any, Any, str]]:
    """対面ディレクトリ内の xlsx / csv から (乗務員ID, 乗務員名, 日時, 種別) イベント列。"""
    events: List[Tuple[Any, Any, Any, str]] = []
    if not taimen_dir.exists():
        return events
    for path in sorted(taimen_dir.iterdir()):
        if path.name.startswith("~") or not path.is_file():
            continue
        suf = path.suffix.lower()
        is_csv = suf == ".csv" or (suf == "" and path.name.lower().endswith("csv"))
        if suf == ".xlsx":
            try:
                wb = load_workbook(path, read_only=False, data_only=True)
                for sheet in wb.worksheets:
                    events.extend(load_taimen_sheet(sheet))
                wb.close()
            except Exception:
                continue
        elif is_csv:
            try:
                events.extend(load_taimen_csv(path))
            except Exception:
                continue
    return events


def _enkaku_dt_indices_from_header(header_cells: List[Any]) -> Tuple[int, int]:
    """遠隔のヘッダー行から日時列（「日時」または「検知日時」）の位置を解決する。
    複数グループ（出庫・中間・帰庫）のうち最初を出庫、最後を帰庫として扱う。
    見つからない場合は従来の固定位置（4, 25）へフォールバック。"""
    idxs = [
        i for i, c in enumerate(header_cells)
        if _normalize_header(str(c) if c is not None else "") in DT_HEADER_NAMES
    ]
    if len(idxs) >= 2:
        return idxs[0], idxs[-1]
    return ENKAKU_COL_INDEX["departure"], ENKAKU_COL_INDEX["return_"]


def load_enkaku_sheet(ws: Worksheet) -> List[Tuple[Any, Any, Any, Any]]:
    """遠隔シートから (社員コード, 社員名, 出庫日時, 帰庫日時) のリスト。"""
    header_cells = [_cell_value(ws, 1, c) for c in range(1, 40)]
    dep_idx, ret_idx = _enkaku_dt_indices_from_header(header_cells)
    data_start = 2
    rows: List[Tuple[Any, Any, Any, Any]] = []
    for r in range(data_start, ws.max_row + 1):
        staff_code = _cell_value(ws, r, ENKAKU_COL_INDEX["staff_code"] + 1)
        staff_name = _cell_value(ws, r, ENKAKU_COL_INDEX["staff_name"] + 1)
        departure = _cell_value(ws, r, dep_idx + 1)
        return_ = _cell_value(ws, r, ret_idx + 1)
        if staff_code is None and staff_name is None:
            continue
        rows.append((staff_code, staff_name, departure, return_))
    return rows


def load_enkaku_csv(path: Path) -> List[Tuple[Any, Any, Any, Any]]:
    """遠隔CSVから (社員コード, 社員名, 出庫日時, 帰庫日時) のリスト。
    日時列はヘッダー名（日時/検知日時）から解決し、不明時は従来の固定位置（4, 25）。"""
    rows = _read_csv_rows(path)
    if len(rows) < 2:
        return []
    dep_idx, ret_idx = _enkaku_dt_indices_from_header(rows[0])
    out: List[Tuple[Any, Any, Any, Any]] = []
    for row in rows[1:]:
        nc = len(row)
        staff_code = row[ENKAKU_COL_INDEX["staff_code"]] if nc > 0 else None
        staff_name = row[ENKAKU_COL_INDEX["staff_name"]] if nc > 1 else None
        departure = row[dep_idx] if nc > dep_idx else None
        return_ = row[ret_idx] if nc > ret_idx else None
        if not staff_code and not staff_name:
            continue
        out.append((staff_code or None, staff_name or None, departure or None, return_ or None))
    return out


def _normalize_plate(v: Any) -> str:
    """車両番号の比較用正規化（全角半角統一・空白除去）。"""
    import unicodedata
    s = unicodedata.normalize("NFKC", str(v or "")).replace(" ", "").replace("　", "")
    return s


def enkaku_vehicles_by_crew(alcohol_dir: Path) -> Dict[str, set]:
    """遠隔（アルキラー）から、乗務員ID正規化 → 検査時に記録された車両番号の集合を返す。
    車両列は各日時列（日時/検知日時）の直後の列とする。日報ダウンロード漏れ検知用。"""
    out: Dict[str, set] = {}
    if not alcohol_dir.exists():
        return out

    def _collect(header_cells: List[Any], get_cell) -> None:
        dt_idxs = [
            i for i, c in enumerate(header_cells)
            if _normalize_header(str(c) if c is not None else "") in DT_HEADER_NAMES
        ]
        if not dt_idxs:
            dt_idxs = [ENKAKU_COL_INDEX["departure"], ENKAKU_COL_INDEX["return_"]]
        veh_idxs = [i + 1 for i in dt_idxs]
        for row_vals in get_cell:
            code = row_vals[ENKAKU_COL_INDEX["staff_code"]] if len(row_vals) > 0 else None
            norm = _normalize_crew_id(code)
            if not norm:
                continue
            for vi in veh_idxs:
                if vi < len(row_vals):
                    plate = _normalize_plate(row_vals[vi])
                    if plate:
                        out.setdefault(norm, set()).add(plate)

    for path in sorted(alcohol_dir.iterdir()):
        if path.name.startswith("~") or not path.is_file():
            continue
        suf = path.suffix.lower()
        is_csv = suf == ".csv" or (suf == "" and path.name.lower().endswith("csv"))
        if suf == ".xlsx":
            try:
                wb = load_workbook(path, read_only=False, data_only=True)
                for sheet in wb.worksheets:
                    header = [_cell_value(sheet, 1, c) for c in range(1, 40)]
                    data = [
                        [_cell_value(sheet, r, c) for c in range(1, 40)]
                        for r in range(2, sheet.max_row + 1)
                    ]
                    _collect(header, data)
                wb.close()
            except Exception:
                continue
        elif is_csv:
            try:
                rows = _read_csv_rows(path)
                if len(rows) >= 2:
                    _collect(rows[0], rows[1:])
            except Exception:
                continue
    return out


def load_enkaku_events(alcohol_dir: Path) -> List[Tuple[Any, Any, Any, str]]:
    """遠隔ディレクトリ内の xlsx / csv から、1行を出庫イベント＋帰庫イベントに展開。"""
    events: List[Tuple[Any, Any, Any, str]] = []
    if not alcohol_dir.exists():
        return events
    for path in sorted(alcohol_dir.iterdir()):
        if path.name.startswith("~") or not path.is_file():
            continue
        suf = path.suffix.lower()
        is_csv = suf == ".csv" or (suf == "" and path.name.lower().endswith("csv"))
        if suf == ".xlsx":
            try:
                wb = load_workbook(path, read_only=False, data_only=True)
                for sheet in wb.worksheets:
                    for (staff_code, staff_name, departure, return_) in load_enkaku_sheet(sheet):
                        if departure is not None and str(departure).strip():
                            events.append((staff_code, staff_name, departure, "出庫"))
                        if return_ is not None and str(return_).strip():
                            events.append((staff_code, staff_name, return_, "帰庫"))
                wb.close()
            except Exception:
                continue
        elif is_csv:
            try:
                for (staff_code, staff_name, departure, return_) in load_enkaku_csv(path):
                    if departure and str(departure).strip():
                        events.append((staff_code, staff_name, departure, "出庫"))
                    if return_ and str(return_).strip():
                        events.append((staff_code, staff_name, return_, "帰庫"))
            except Exception:
                continue
    return events


# 統合イベント = (乗務員ID, 乗務員名, 日時, 種別). 日時は元の型のまま（ソート時のみ datetime 使用）
AlcoholEvent = Tuple[Any, Any, Any, str]


def integrate_alcohol(
    taimen_dir: Path,
    alcohol_dir: Path,
) -> List[AlcoholEvent]:
    """
    対面・遠隔を紐づけず、乗務員ID順・日時昇順で並べたイベント列を返す。
    各要素は (乗務員ID, 乗務員名, 日時, 種別)。日時が解釈できない行は除外する。
    """
    events: List[AlcoholEvent] = []
    events.extend(load_taimen_events(taimen_dir))
    events.extend(load_enkaku_events(alcohol_dir))

    # 日時でソートするため datetime に変換できるものだけ残し、乗務員ID・日時でソート
    def sort_key(e: AlcoholEvent) -> Tuple[Any, datetime]:
        uid, _, dt, _ = e
        t = _to_datetime(dt)
        return (str(uid) if uid is not None else "", t or datetime.max)

    events = [e for e in events if _to_datetime(e[2]) is not None]

    # 重複除去: アルキラーNEXの日別シートは日をまたぐ運行の記録を
    # 前日・当日の両方のシートに出力するため、同一の検査が2回読み込まれる。
    # 乗務員ID（正規化）・日時・種別が完全一致するイベントは1件にまとめる。
    seen: set = set()
    deduped: List[AlcoholEvent] = []
    for e in events:
        uid, _name, dt, kind = e
        key = (_normalize_crew_id(uid), _to_datetime(dt), kind)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(e)
    events = deduped

    events.sort(key=sort_key)
    return events


def _normalize_crew_id(uid: Any) -> str:
    """乗務員ID比較用: 前後空白除去し、先頭の0を除いた数字に正規化。'00002433' と '2433' を一致させる。"""
    s = str(uid).strip() if uid is not None else ""
    if not s:
        return ""
    s = s.lstrip("0") or "0"  # 全部0の場合は "0"
    return s


def match_alcohol_for_run(
    alcohol_events: List[AlcoholEvent],
    crew_id: Any,
    digitaco_out_dt: Any,
    digitaco_in_dt: Any,
    margin_minutes: int = 120,
) -> Tuple[Optional[datetime], Optional[datetime]]:
    """
    デジタコ1運行に対するアルコール突合。出庫同士・帰庫同士のみ。±margin_minutes 以内で最も近いものを採用。
    マッチしなければ None。戻り値は (出庫日時 or None, 帰庫日時 or None)。datetime で返す。
    """
    crew_norm = _normalize_crew_id(crew_id)
    out_dt = _to_datetime(digitaco_out_dt)
    in_dt = _to_datetime(digitaco_in_dt)

    same_crew = [e for e in alcohol_events if _normalize_crew_id(e[0]) == crew_norm]
    out_events = [(e[2], _to_datetime(e[2])) for e in same_crew if e[3] == "出庫"]
    in_events = [(e[2], _to_datetime(e[2])) for e in same_crew if e[3] == "帰庫"]

    def best_in_window(events: List[Tuple[Any, datetime]], center: Optional[datetime]) -> Optional[datetime]:
        if center is None:
            return None
        low = center - timedelta(minutes=margin_minutes)
        high = center + timedelta(minutes=margin_minutes)
        in_range = [(orig, t) for orig, t in events if t is not None and low <= t <= high]
        if not in_range:
            return None
        in_range.sort(key=lambda x: abs((x[1] - center).total_seconds()))
        return in_range[0][1]

    matched_out = best_in_window(out_events, out_dt)
    matched_in = best_in_window(in_events, in_dt)
    return (matched_out, matched_in)


def format_dt_for_excel(dt: Optional[datetime]) -> Optional[str]:
    """Excel 出力用に datetime をデジタコと同じ文字列に。None はそのまま。"""
    if dt is None:
        return None
    return dt.strftime(DT_FMT)


# 同乗者用: 出庫のみ・帰庫のみ・24時間以内の帰庫採用
ALCOHOL_RUN_24H_HOURS = 24


def alcohol_runs_by_crew(
    alcohol_events: List[AlcoholEvent],
) -> Dict[str, List[Dict[str, str]]]:
    """
    乗務員ID（正規化）ごとに、1運行を時系列でまとめる。
    - 出庫 → 24時間以内の帰庫のうち最も近い1本を採用 → 1運行。なければ出庫のみ1運行。
    - 出庫の次が出庫 → 出庫のみ1運行。
    - 帰庫 → 出庫: 帰庫のみ1運行、出庫のみ1運行。
    - 帰庫 → 帰庫: それぞれ帰庫のみ1運行。
    戻り値: { 乗務員ID正規化: [ {"出庫日時": "..." or "", "帰庫日時": "..." or ""}, ... ] }
    """
    from collections import defaultdict

    by_crew: Dict[str, List[Tuple[datetime, str]]] = defaultdict(list)
    for e in alcohol_events:
        uid, _, dt_val, typ = e
        t = _to_datetime(dt_val)
        if t is None:
            continue
        crew_norm = _normalize_crew_id(uid)
        by_crew[crew_norm].append((t, typ))

    out: Dict[str, List[Dict[str, str]]] = {}
    for crew_norm, events in by_crew.items():
        events.sort(key=lambda x: x[0])
        runs: List[Dict[str, str]] = []
        used_in: set = set()  # index of 帰庫 already paired with an 出庫

        i = 0
        while i < len(events):
            t_i, typ_i = events[i][0], events[i][1]
            if typ_i == "帰庫":
                # すでに直前の出庫とペアにした帰庫は「帰庫のみ」で重複出力しない
                if i not in used_in:
                    runs.append({"出庫日時": "", "帰庫日時": t_i.strftime(DT_FMT)})
                i += 1
                continue
            if typ_i == "出庫":
                out_dt = t_i
                window_end = out_dt + timedelta(hours=ALCOHOL_RUN_24H_HOURS)
                candidates = [
                    (j, events[j][0])
                    for j in range(i + 1, len(events))
                    if events[j][1] == "帰庫"
                    and j not in used_in
                    and out_dt <= events[j][0] <= window_end
                ]
                if candidates:
                    candidates.sort(key=lambda x: abs((x[1] - out_dt).total_seconds()))
                    j_best, in_dt = candidates[0]
                    used_in.add(j_best)
                    runs.append({
                        "出庫日時": out_dt.strftime(DT_FMT),
                        "帰庫日時": in_dt.strftime(DT_FMT),
                    })
                else:
                    runs.append({"出庫日時": out_dt.strftime(DT_FMT), "帰庫日時": ""})
                i += 1
                continue
            i += 1
        out[crew_norm] = runs
    return out


def alcohol_only_crew_list(
    alcohol_events: List[AlcoholEvent],
    crew_ids_in_digitaco: set,
) -> List[Dict[str, Any]]:
    """
    アルコールにはいるがデジタコの運行に1件も出てこない乗務員のリストを返す。
    各要素: { "乗務員ID正規化", "乗務員ID", "乗務員名", "runs": [ {"出庫日時","帰庫日時"}, ... ] }
    """
    runs_by_crew = alcohol_runs_by_crew(alcohol_events)
    crew_first: Dict[str, Tuple[Any, Any]] = {}  # crew_norm -> (乗務員ID, 乗務員名)
    for e in alcohol_events:
        uid, name, _, _ = e
        crew_norm = _normalize_crew_id(uid)
        if crew_norm not in crew_first:
            crew_first[crew_norm] = (uid, name)
    out: List[Dict[str, Any]] = []
    for crew_norm, runs in runs_by_crew.items():
        if crew_norm in crew_ids_in_digitaco:
            continue
        if not runs:
            continue
        uid, name = crew_first.get(crew_norm, (None, None))
        out.append({
            "乗務員ID正規化": crew_norm,
            "乗務員ID": uid,
            "乗務員名": name,
            "runs": runs,
        })
    return out


def write_integrated_excel(events: List[AlcoholEvent], out_path: Path) -> None:
    """統合イベントを Excel で出力。A=乗務員ID, B=乗務員名, C=日時, D=種別。"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("no active sheet")
    ws.title = "アルコール統合"
    ws.append(["乗務員ID", "乗務員名", "日時", "種別"])
    for e in events:
        ws.append(list(e))
    wb.save(out_path)
