from __future__ import annotations

import csv
import json
import logging
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timedelta, time
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional

from .alcohol_integration import (
    integrate_alcohol,
    match_alcohol_for_run,
    format_dt_for_excel,
    alcohol_runs_by_crew as get_alcohol_runs_by_crew,
)

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font

# 運転時間比較ログ用ファイル（_merge_runs 直後 vs rows_from_run_states 再計算後）
_DRIVE_COMPARE_LOG_PATH = Path(__file__).resolve().parents[1] / "work" / "drive_compare.log"
_drive_compare_logger: Optional[logging.Logger] = None


def _get_drive_compare_logger() -> logging.Logger:
    global _drive_compare_logger
    if _drive_compare_logger is not None:
        return _drive_compare_logger
    _drive_compare_logger = logging.getLogger("engine.pipeline.drive_compare")
    _drive_compare_logger.setLevel(logging.INFO)
    _DRIVE_COMPARE_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    h = logging.FileHandler(_DRIVE_COMPARE_LOG_PATH, encoding="utf-8")
    h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    _drive_compare_logger.addHandler(h)
    return _drive_compare_logger


# =========================
# Types
# =========================

@dataclass
class PipelineResult:
    excel_path: Path
    log_path: Path
    skipped_path: Path
    error_count: int
    warn_count: int
    manual_input_required: bool = False
    merge_decision_required: bool = False  # 同一乗務員・複数運行を1つにするか質問
    merge_groups: Optional[List[Dict[str, Any]]] = None  # [{ rowIndices: [i,j], 運行IDs: [...], 乗務員ID, 乗務員名 }, ...]
    run_states: Optional[List[Dict[str, Any]]] = None  # 手入力完了時に再計算する用
    pending_rows: Optional[List[Dict[str, Any]]] = None  # 出庫・帰庫が未取得の行一覧（画面表示用）
    headers: Optional[List[str]] = None  # 手入力完了時に Excel を書く用
    alcohol_runs_by_crew: Optional[Dict[str, List[Dict[str, str]]]] = None  # 同乗者用: 乗務員ID正規化 -> [{"出庫日時","帰庫日時"}, ...]


# =========================
# Utilities
# =========================

def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _nfkc(s: str) -> str:
    return unicodedata.normalize("NFKC", s)


def _clean_for_regex(s: str) -> str:
    s = _nfkc(s)
    s = re.sub(r"[\x00-\x1F]+", " ", s)
    return s


def _parse_dt(s: str) -> datetime:
    return datetime.strptime(s, "%Y/%m/%d %H:%M")


def _row_to_dt(v: Any) -> Optional[datetime]:
    """行の日時項目を datetime に。None/空/解釈不可なら None。全角数字・記号は NFKC で半角に正規化してから解釈。"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = _nfkc(str(v).strip())
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y/%m/%d %H:%M")
    except ValueError:
        pass
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _minutes_between(a: datetime, b: datetime) -> int:
    # 分単位で「差分」を出す。秒は無視（PDF時刻は分精度なのでこれで十分）
    return int((a - b).total_seconds() // 60)


def _split_day_night_minutes(start: datetime, end: datetime) -> Tuple[int, int]:
    """
    Day: 05:00-22:00
    Night: 22:00-05:00
    1分単位で分割
    """
    if end <= start:
        return 0, 0

    day = 0
    night = 0
    cur = start
    while cur < end:
        nxt = min(end, cur + timedelta(minutes=1))
        t = cur.time()
        is_day = time(5, 0) <= t < time(22, 0)
        if is_day:
            day += 1
        else:
            night += 1
        cur = nxt
    return day, night


def _apply_regex(text: str, pattern: str, group: int = 1) -> Optional[str]:
    m = re.search(pattern, text, re.MULTILINE)
    return m.group(group).strip() if m else None


def _extract_out_in_dt_pair(text: str, out_pattern: str, in_pattern: str) -> Tuple[Optional[str], Optional[str]]:
    """
    ブロック内に複数の「出庫時刻」「帰庫時刻」がある場合、同一運行の対を返す。
    「最初の出庫」と組にする帰庫は、その出庫より後に現れる帰庫のうち、
    拘束時間（出庫→帰庫の分）が最短になるものを採用する。
    （正しいペアは同一運行で約数～十数時間、誤ったペアは他運行の帰庫で20時間超になるため）
    """
    outs: List[Tuple[int, str]] = []  # (start_pos, value)
    for m in re.finditer(out_pattern, text, re.MULTILINE):
        outs.append((m.start(), m.group(1).strip()))
    ins: List[Tuple[int, str]] = []
    for m in re.finditer(in_pattern, text, re.MULTILINE):
        ins.append((m.start(), m.group(1).strip()))
    if not outs:
        return None, None
    first_out_pos, first_out_val = outs[0]
    if not ins:
        return first_out_val, None
    # 最初の出庫より後に出てくる帰庫のうち、拘束時間が最短（正の分）のものを採用
    best_in_val: Optional[str] = None
    best_minutes: Optional[int] = None
    out_dt = _row_to_dt(first_out_val)
    if out_dt is None:
        return first_out_val, ins[0][1] if ins else None
    for _pos, in_val in ins:
        if _pos <= first_out_pos:
            continue
        in_dt = _row_to_dt(in_val)
        if in_dt is None or in_dt <= out_dt:
            continue
        minutes = _minutes_between(in_dt, out_dt)
        if best_minutes is None or minutes < best_minutes:
            best_minutes = minutes
            best_in_val = in_val
    if best_in_val is not None:
        return first_out_val, best_in_val
    return first_out_val, None


def _is_time_serial_col(header_name: str) -> bool:
    """
    Excel上で h:mm 表示にしたい列（中身は「分」で計算→書き込み時に time serial へ変換）
    """
    return header_name in {
        "拘束時間_分割前",
        "拘束時間_昼_分割前",
        "拘束時間_夜_分割前",
        "拘束時間_分割後",
        "拘束時間_昼_分割後",
        "拘束時間_夜_分割後",
        "労働時間_分割前",
        "労働時間_昼_分割前",
        "労働時間_夜_分割前",
        "労働時間_分割後",
        "労働時間_昼_分割後",
        "労働時間_夜_分割後",
        "運転時間",
        "待機時間",
        "荷積時間",
        "荷卸時間",
        "作業時間",
        "休憩時間_分割前",
        "休憩時間_昼_分割前",
        "休憩時間_夜_分割前",
        "休憩時間_分割後",
        "休憩時間_昼_分割後",
        "休憩時間_夜_分割後",
        "休息時間",
        "休息時間_昼",
        "休息時間_夜",
    }


def _minutes_to_excel_time_serial(minutes: Optional[int]) -> Optional[float]:
    if minutes is None:
        return None
    if minutes < 0:
        minutes = 0
    return minutes / 1440.0  # 1日=1440分


# =========================
# PDF read
# =========================

def _read_pdf_text(pdf_path: Path) -> Tuple[str, str]:
    texts: List[str] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for p in pdf.pages:
            texts.append(p.extract_text() or "")
    raw = "\n".join(texts)
    cleaned = _clean_for_regex(raw)
    return raw, cleaned


def _split_raw_by_runs(raw: str, report_id_regex: str) -> List[str]:
    """
    1PDF内に複数運行がある場合、運行IDが現れる行で分割する。
    運行IDが1件も無い場合は元のrawを1ブロックとして返す（①⑤互換）。
    """
    lines = raw.splitlines()
    if not lines:
        return [raw]
    he = re.compile(report_id_regex)
    indices: List[int] = []
    for i, line in enumerate(lines):
        if he.search(_nfkc(line)):
            indices.append(i)
    if not indices:
        return [raw]
    blocks: List[str] = []
    for k in range(len(indices)):
        start = indices[k]
        end = indices[k + 1] if k + 1 < len(indices) else len(lines)
        block_lines = lines[start:end]
        blocks.append("\n".join(block_lines))
    return blocks


# =========================
# Header extract
# =========================

def _load_preset(preset_path: Path) -> Dict[str, Any]:
    return json.loads(preset_path.read_text(encoding="utf-8"))


def _extract_header_fields(cleaned_text: str, device: str, preset: Dict[str, Any]) -> Dict[str, Any]:
    """
    埋める対象（現時点）:
    - 運行ID / 運行日
    - 乗務員名 / 乗務員ID
    - 車両番号 / 車両ID（括弧内）
    - 所属営業所 / 営業所ID（括弧内）
    - 出庫日時 / 帰庫日時（ヘッダー）
    - 出庫メーター / 帰庫メーター（小数あり）
    """
    fields: Dict[str, Any] = {}
    he: Dict[str, str] = (preset.get("header_extract") or {})

    if device == "mimamori":
        rid = he.get("report_id_regex") or r"(ID-\d+)"
        run = he.get("run_date_regex") or r"(運行日付|運行日)\s*[:：]?\s*(\d{4}/\d{1,2}/\d{1,2})"
        drv = he.get("driver_regex") or r"乗務員名\s*[:：]\s*(.+?)\s*\((\d+)\)"
        veh = he.get("vehicle_regex") or r"車両名称\s*[:：]\s*(.+?)\s*\((\d+)\)"
        off = he.get("office_regex") or r"所\s*属\s*[:：]\s*(.+?)\s*\((\d+)\)"
        mout = he.get("meter_out_regex") or r"出庫時刻\s*[:：]\s*\d{4}/\d{1,2}/\d{1,2}\s*\d{2}:\d{2}\s*([0-9]+(?:\.[0-9]+)?)\s*km"
        min_ = he.get("meter_in_regex") or r"帰庫時刻\s*[:：]\s*\d{4}/\d{1,2}/\d{1,2}\s*\d{2}:\d{2}\s*([0-9]+(?:\.[0-9]+)?)\s*km"
        outdt = he.get("header_out_dt_regex") or r"出庫時刻\s*[:：]\s*(\d{4}/\d{1,2}/\d{1,2}\s*\d{2}:\d{2})"
        indt = he.get("header_in_dt_regex") or r"帰庫時刻\s*[:：]\s*(\d{4}/\d{1,2}/\d{1,2}\s*\d{2}:\d{2})"

        fields["運行ID"] = _apply_regex(cleaned_text, rid, 1)
        fields["運行日"] = _apply_regex(cleaned_text, run, 2)

        m = re.search(drv, cleaned_text)
        if m:
            fields["乗務員名"] = m.group(1).strip()
            fields["乗務員ID"] = m.group(2).strip()

        m = re.search(veh, cleaned_text)
        if m:
            fields["車両番号"] = m.group(1).strip()
            fields["車両ID"] = m.group(2).strip()

        m = re.search(off, cleaned_text)
        if m:
            fields["所属営業所"] = m.group(1).strip()
            fields["営業所ID"] = m.group(2).strip()

        fields["出庫メーター"] = _apply_regex(cleaned_text, mout, 1)
        fields["帰庫メーター"] = _apply_regex(cleaned_text, min_, 1)
        out_dt_val, in_dt_val = _extract_out_in_dt_pair(cleaned_text, outdt, indt)
        fields["出庫日時"] = out_dt_val
        fields["帰庫日時"] = in_dt_val

        # 安全点数
        score_safe = he.get("score_safe_regex")
        if score_safe:
            fields["安全点数"] = _apply_regex(cleaned_text, score_safe, 1)

    else:
        # telecom
        fields["運行ID"] = _apply_regex(cleaned_text, r"(ID-\d+)", 1)
        fields["運行日"] = _apply_regex(cleaned_text, r"運行日付[:：]?\s*(\d{4}/\d{1,2}/\d{1,2})", 1)

        m = re.search(r"乗務員名[:：]\s*(.+?)\s*\((\d+)\)", cleaned_text)
        if m:
            fields["乗務員名"] = m.group(1).strip()
            fields["乗務員ID"] = m.group(2).strip()

        m = re.search(r"車両名称[:：]\s*(.+?)\s*\((\d+)\)", cleaned_text)
        if m:
            fields["車両番号"] = m.group(1).strip()
            fields["車両ID"] = m.group(2).strip()

        m = re.search(r"所\s*属\s*[:：]\s*(.+?)\s*\((\d+)\)", cleaned_text)
        if m:
            fields["所属営業所"] = m.group(1).strip()
            fields["営業所ID"] = m.group(2).strip()

        fields["出庫メーター"] = _apply_regex(
            cleaned_text,
            r"出庫時刻[:：]\s*\d{4}/\d{1,2}/\d{1,2}\s*\d{2}:\d{2}\s*([0-9]+(?:\.[0-9]+)?)\s*km",
            1,
        )
        fields["帰庫メーター"] = _apply_regex(
            cleaned_text,
            r"帰庫時刻[:：]\s*\d{4}/\d{1,2}/\d{1,2}\s*\d{2}:\d{2}\s*([0-9]+(?:\.[0-9]+)?)\s*km",
            1,
        )
        _out_re = r"出庫時刻[:：]\s*(\d{4}/\d{1,2}/\d{1,2}\s*\d{2}:\d{2})"
        _in_re = r"帰庫時刻[:：]\s*(\d{4}/\d{1,2}/\d{1,2}\s*\d{2}:\d{2})"
        out_dt_val, in_dt_val = _extract_out_in_dt_pair(cleaned_text, _out_re, _in_re)
        fields["出庫日時"] = out_dt_val
        fields["帰庫日時"] = in_dt_val

        # 安全点数
        score_safe = he.get("score_safe_regex")
        if score_safe:
            fields["安全点数"] = _apply_regex(cleaned_text, score_safe, 1)

    for k in ("出庫メーター", "帰庫メーター"):
        if fields.get(k) is not None:
            try:
                fields[k] = round(float(fields[k]), 1)  # 小数点第1位まで
            except Exception:
                pass

    # 走行状態の次の時刻（H:MM）を分で取得。1運行1つなのでここで取れば運転時間として合算可能
    drive_re = (preset.get("header_extract") or {}).get("drive_time_regex")
    if not drive_re:
        drive_re = r"走行状態\s*[:：]?\s*(\d{1,2}):(\d{2})"
    m = re.search(drive_re, cleaned_text)
    if m:
        try:
            h, mn = int(m.group(1)), int(m.group(2))
            if 0 <= h < 24 and 0 <= mn < 60:
                fields["走行状態_分"] = h * 60 + mn
        except (ValueError, IndexError):
            pass

    return fields


# =========================
# Detail extract (items)
# =========================

def _extract_detail_rows(raw_text: str) -> List[Dict[str, Any]]:
    """
    明細の1行から最小で取りたいもの:
    - item番号
    - 作業（休憩 / 待機 / 出庫 / 帰庫 ...）
    - 到着 HH:MM（無い場合あり）
    - 出発 HH:MM（無い場合あり）

    mimamoriは左右2カラムが同一行に混ざることがあるので、同一行内で item開始が複数なら分割抽出。
    """
    rows: List[Dict[str, Any]] = []

    for line in raw_text.splitlines():
        line = _nfkc(line).strip()
        if not line:
            continue
        if not re.match(r"^\d+\s", line):
            continue
        if re.match(r"^\d{1,2}月\d{1,2}日", line):
            continue

        starts: List[int] = []
        for m in re.finditer(r"(?:(?<=^)|(?<=\s))(\d{1,3})\s+([^\d\s]\S*)", line):
            starts.append(m.start(1))
        if not starts:
            continue
        starts = sorted(set(starts))

        segments: List[str] = []
        for i, s in enumerate(starts):
            e = starts[i + 1] if i + 1 < len(starts) else len(line)
            seg = line[s:e].strip()
            segments.append(seg)

        for seg in segments:
            toks = seg.split()
            if len(toks) < 2:
                continue
            try:
                item = int(toks[0])
            except Exception:
                continue
            task = toks[1]
            times = re.findall(r"\b\d{2}:\d{2}\b", seg)
            arr = times[0] if len(times) >= 1 else None
            dep = times[1] if len(times) >= 2 else None
            rows.append({"item": item, "task": task, "arrival": arr, "depart": dep, "raw": seg})

    seen = set()
    out: List[Dict[str, Any]] = []
    for r in rows:
        key = (r["item"], r["task"], r["arrival"], r["depart"])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def _build_datetime_sequence(out_dt: datetime, detail_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    ルール: 前行より時刻が戻ったら翌日扱い（rollover +1 day）
    """
    prev = out_dt
    seq: List[Dict[str, Any]] = []

    def make_dt(hhmm: str, ref: datetime) -> datetime:
        h, m = map(int, hhmm.split(":"))
        candidate = datetime.combine(ref.date(), time(h, m))
        if candidate < ref:
            candidate += timedelta(days=1)
        return candidate

    for r in detail_rows:
        arr_dt = None
        dep_dt = None

        if r["arrival"]:
            arr_dt = make_dt(r["arrival"], prev)
            prev = arr_dt

        if r["depart"]:
            dep_dt = make_dt(r["depart"], prev)
            prev = dep_dt

        seq.append({**r, "arr_dt": arr_dt, "dep_dt": dep_dt})

    return seq


# =========================
# Metrics compute (ALL INTERNAL = minutes)
# =========================

def _compute_metrics(header: Dict[str, Any], detail_rows: List[Dict[str, Any]], logs: List[Dict[str, Any]], ctx: Dict[str, Any], preset: Dict[str, Any]) -> Dict[str, Any]:
    out_dt_s = header.get("出庫日時")
    in_dt_s = header.get("帰庫日時")

    if not out_dt_s or not in_dt_s:
        logs.append({
            **ctx,
            "level": "ERROR",
            "category": "HEADER_MISSING",
            "field_name": "出庫日時/帰庫日時",
            "value_candidates": "",
            "message": "ヘッダーから出庫日時/帰庫日時が取れないため集計不可",
        })
        return {}

    out_dt = _parse_dt(out_dt_s)
    in_dt = _parse_dt(in_dt_s)

    def _detail_sort_key(r: Dict[str, Any]):
        if r.get("_merge_seq") is not None:
            return (0, int(r["_merge_seq"]))
        item = r.get("item")
        item = item if isinstance(item, int) else 10**9
        return (1, item, r.get("arrival") or "99:99", r.get("depart") or "99:99")

    detail_rows_sorted = sorted(detail_rows, key=_detail_sort_key)
    seq = _build_datetime_sequence(out_dt, detail_rows_sorted)

    # 総走行距離 = 帰庫 - 出庫
    dist = None
    try:
        if header.get("出庫メーター") is not None and header.get("帰庫メーター") is not None:
            dist = float(header["帰庫メーター"]) - float(header["出庫メーター"])
    except Exception:
        dist = None

    # 拘束時間（分）
    bind_min = _minutes_between(in_dt, out_dt)
    # 拘束時間の昼/夜内訳（5:00-22:00=昼、22:00-5:00=夜）
    bind_day, bind_night = _split_day_night_minutes(out_dt, in_dt)

    # 運転時間（分）：PDFの「走行状態」の次の時刻を優先。無い場合は明細シーケンスから算出
    drive_min: Optional[int] = None
    if header.get("走行状態_分") is not None:
        try:
            v = header["走行状態_分"]
            drive_min = int(v) if isinstance(v, (int, float)) else int(float(str(v).strip()))
            if drive_min < 0:
                drive_min = 0
        except (ValueError, TypeError):
            pass
    if drive_min is None:
        drive_min = 0
        prev_depart: Optional[datetime] = out_dt

        for s in seq:
            if s["arr_dt"] is not None and prev_depart is not None:
                if s["arr_dt"] >= prev_depart:
                    drive_min += _minutes_between(s["arr_dt"], prev_depart)
            if s["task"] == "出庫" and s["arr_dt"] is not None and s["dep_dt"] is None:
                prev_depart = s["arr_dt"]
            elif s["task"] == "帰庫" and s["arr_dt"] is not None and s["dep_dt"] is None:
                prev_depart = s["arr_dt"]
            elif s["dep_dt"] is not None:
                prev_depart = s["dep_dt"]

        if prev_depart is not None and in_dt >= prev_depart:
            drive_min += _minutes_between(in_dt, prev_depart)

    # 待機時間（分）
    wait_min = 0
    for s in seq:
        if s["task"] == "待機" and s["arr_dt"] is not None and s["dep_dt"] is not None:
            wait_min += _minutes_between(s["dep_dt"], s["arr_dt"])

    # 荷積時間・荷卸時間・作業時間（分）：プリセットの task_names に従い作業明細を集計
    task_names = preset.get("task_names") or {}
    niomi_list = task_names.get("荷積") or ["荷積"]
    niose_list = task_names.get("荷卸") or ["荷卸"]
    gyomu_only_list = task_names.get("作業時間のみ") or []
    niomi_min = 0
    niose_min = 0
    gyomu_only_min = 0
    for s in seq:
        if s["arr_dt"] is not None and s["dep_dt"] is not None:
            dur = _minutes_between(s["dep_dt"], s["arr_dt"])
            task = s["task"]
            if task in niomi_list:
                niomi_min += dur
            if task in niose_list:
                niose_min += dur
            if task in gyomu_only_list:
                gyomu_only_min += dur
    work_total_min = niomi_min + niose_min + gyomu_only_min  # 作業時間 = 荷積+荷卸+作業時間のみ

    # 休憩（対象は 作業=休憩 のみ）
    breaks: List[Dict[str, Any]] = []
    for s in seq:
        if s["task"] == "休憩" and s["arr_dt"] is not None and s["dep_dt"] is not None:
            dur = _minutes_between(s["dep_dt"], s["arr_dt"])
            breaks.append({"start": s["arr_dt"], "end": s["dep_dt"], "dur": dur})

    # 休息候補（>=180分）を上位2回だけ休息扱い
    candidates = [b for b in breaks if b["dur"] >= 180]
    selected = sorted(candidates, key=lambda x: (-x["dur"], x["start"]))[:2]
    selected_set = {(b["start"], b["end"]) for b in selected}

    break_total = break_day = break_night = 0
    rest_total = rest_day = rest_night = 0
    rest_splits: List[Dict[str, Any]] = []

    for b in breaks:
        d, n = _split_day_night_minutes(b["start"], b["end"])
        if (b["start"], b["end"]) in selected_set:
            rest_total += b["dur"]
            rest_day += d
            rest_night += n
            rest_splits.append(b)
        else:
            break_total += b["dur"]
            break_day += d
            break_night += n

    rest_splits = sorted(rest_splits, key=lambda x: (-x["dur"], x["start"]))[:2]

    # 労働時間 = 拘束時間 - 休憩時間（分割前のみ。分割後は空欄）
    work_bind_min = max(0, bind_min - break_total)
    work_bind_day = max(0, bind_day - break_day)
    work_bind_night = max(0, bind_night - break_night)

    out: Dict[str, Any] = {
        "総走行距離": dist,

        # ★ここから全部「分(int)」で保持（Excel表示だけ後で h:mm にする）
        "拘束時間_分割前": int(bind_min),
        "拘束時間_昼_分割前": int(bind_day),
        "拘束時間_夜_分割前": int(bind_night),
        # 拘束時間_分割後 / 拘束時間_昼_分割後 / 拘束時間_夜_分割後 → 何も入れない（ヘッダーのみ）

        "労働時間_分割前": int(work_bind_min),
        "労働時間_昼_分割前": int(work_bind_day),
        "労働時間_夜_分割前": int(work_bind_night),
        # 労働時間_分割後 / 労働時間_昼_分割後 / 労働時間_夜_分割後 → 何も入れない（ヘッダーのみ）

        "運転時間": int(drive_min),
        "待機時間": int(wait_min),
        "荷積時間": int(niomi_min),
        "荷卸時間": int(niose_min),
        "作業時間": int(work_total_min),

        "休憩時間_分割前": int(break_total),
        "休憩時間_昼_分割前": int(break_day),
        "休憩時間_夜_分割前": int(break_night),

        "休息時間": int(rest_total),
        "休息時間_昼": int(rest_day),
        "休息時間_夜": int(rest_night),
    }

    for idx, b in enumerate(rest_splits, start=1):
        out[f"分割開始{idx}"] = b["start"]
        out[f"分割終了{idx}"] = b["end"]
        out[f"分割{idx}_作業時間_分"] = int(b["dur"])  # ★この列は“分”のまま

    return out


def _apply_merged_drive_override(row: Dict[str, Any], merged_header: Dict[str, Any]) -> None:
    """統合行の運転時間を _merge_runs 直後の値に戻す。再計算で膨張するのを防ぐ。"""
    v = merged_header.get("_merged_drive_min_initial")
    if v is None:
        return
    try:
        n = int(v)
        if n >= 0:
            row["運転時間"] = n
    except (TypeError, ValueError):
        pass
    row.pop("_merged_drive_min_initial", None)


# =========================
# 手入力完了: 入力値で再計算して Excel 出力
# =========================

def complete_manual_input(
    run_states: List[Dict[str, Any]],
    headers: List[str],
    entries: List[Dict[str, Any]],
    preset_path: Path,
    device: str,
    excel_path: Path,
) -> None:
    """
    手入力された出庫日時・帰庫日時で run_states を更新し、全行の metrics を再計算して Excel を書く。
    entries は [ {"rowIndex": 0, "出庫日時": "...", "帰庫日時": "..." }, ... ]
    同乗者行は {"rowIndex": i, "driverRowIndex": j, "出庫日時": "...", "帰庫日時": "..." } を送る。
    j 行目のデジタコデータを流用し、乗務員ID・乗務員名は i 行目そのまま、出庫・帰庫は入力値を使う。
    """
    preset = _load_preset(preset_path)
    entries_by_index = {int(e["rowIndex"]): e for e in entries}
    rows: List[Dict[str, Any]] = []
    for i, run in enumerate(run_states):
        e = entries_by_index.get(i)
        if e is not None and e.get("driverRowIndex") is not None:
            driver_idx = int(e["driverRowIndex"])
            if 0 <= driver_idx < len(run_states):
                driver_run = run_states[driver_idx]
                merged_header = dict(driver_run["merged_header"])
                merged_details = list(driver_run["merged_details"])
                merged_header["乗務員ID"] = run["merged_header"].get("乗務員ID")
                merged_header["乗務員名"] = run["merged_header"].get("乗務員名")
                merged_header["出庫日時"] = (e.get("出庫日時") or "").strip() or None
                merged_header["帰庫日時"] = (e.get("帰庫日時") or "").strip() or None
            else:
                merged_header = dict(run["merged_header"])
                merged_header["出庫日時"] = (e.get("出庫日時") or "").strip() or None
                merged_header["帰庫日時"] = (e.get("帰庫日時") or "").strip() or None
                merged_details = run["merged_details"]
            ctx = {"timestamp": "", "company": "", "device_type": device, "report_id": "", "pdf_filename": "", "level": "", "category": "", "field_name": "", "value_candidates": "", "message": ""}
            metrics = _compute_metrics(merged_header, merged_details, [], ctx, preset)
            row = {**merged_header, **metrics}
            _apply_merged_drive_override(row, merged_header)
        else:
            merged_header = dict(run["merged_header"])
            if e is not None:
                merged_header["出庫日時"] = (e.get("出庫日時") or "").strip() or None
                merged_header["帰庫日時"] = (e.get("帰庫日時") or "").strip() or None
            merged_details = run["merged_details"]
            if run.get("merged_row") is not None:
                row = dict(run["merged_row"])
                if e is not None:
                    row["出庫日時"] = (e.get("出庫日時") or "").strip() or row.get("出庫日時")
                    row["帰庫日時"] = (e.get("帰庫日時") or "").strip() or row.get("帰庫日時")
            else:
                ctx = {"timestamp": "", "company": "", "device_type": device, "report_id": "", "pdf_filename": "", "level": "", "category": "", "field_name": "", "value_candidates": "", "message": ""}
                metrics = _compute_metrics(merged_header, merged_details, [], ctx, preset)
                row = {**merged_header, **metrics}
                _apply_merged_drive_override(row, merged_header)
        if device in ("telecom", "mimamori") and row.get("運行ID"):
            rid = str(row["運行ID"])
            if rid.startswith("ID-"):
                row["運行ID"] = rid[3:]
        rows.append(row)
    _write_excel(headers, rows, excel_path)


# =========================
# Writers
# =========================

def apply_merge_decision(
    run_states: List[Dict[str, Any]],
    headers: List[str],
    merge_groups: List[Dict[str, Any]],
    merge_choices: List[bool],
    preset_path: Path,
    device: str,
    run_date_choices: Optional[List[int]] = None,
    *,
    merge_sets: Optional[List[List[List[int]]]] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    ユーザーの「まとめる運行」に従い、該当する運行を統合する。
    merge_sets が渡されればそれを使用。無い場合は merge_choices で従来通り（merge_choices[i]=True でグループ全体を統合）。
    merge_sets[gi] = グループ gi の統合セットのリスト。各セットは rowIndex のリスト。len>=2 のセットを1運行に統合する。
    run_date_choices[gi] はグループ gi の統合時に採用する運行日のインデックス（統合セット内の何番目）。省略時は先頭。
    戻り値: (new_run_states, new_rows).
    """
    if merge_sets is None:
        merge_sets = []
        for gi, g in enumerate(merge_groups):
            indices = g.get("rowIndices") or []
            if merge_choices and gi < len(merge_choices) and merge_choices[gi] and len(indices) >= 2:
                merge_sets.append([indices])
            else:
                merge_sets.append([[idx] for idx in indices])
    preset = _load_preset(preset_path)
    ctx = {"timestamp": "", "company": "", "device_type": device, "report_id": "", "pdf_filename": "", "level": "", "category": "", "field_name": "", "value_candidates": "", "message": ""}

    def row_from_run_state(rs: Dict[str, Any]) -> Dict[str, Any]:
        merged_header = dict(rs.get("merged_header") or {})
        if rs.get("merged_row"):
            return dict(rs["merged_row"])
        merged_details = rs.get("merged_details") or []
        metrics = _compute_metrics(merged_header, merged_details, [], ctx, preset)
        row = {**merged_header, **metrics}
        if device in ("telecom", "mimamori") and row.get("運行ID"):
            rid = str(row["運行ID"])
            if rid.startswith("ID-"):
                row["運行ID"] = rid[3:]
        return row

    rows = [row_from_run_state(rs) for rs in run_states]

    # merge_sets の overlap を除去：同じ rowIndex が複数 set に入っていると同一運行が二重計上されるため、先に出た set に属するインデックスは後続から除外する
    used_indices: set = set()
    effective_sets: List[Tuple[int, List[int]]] = []
    for gi, sets in enumerate(merge_sets):
        for s in sets:
            unique_s = sorted(set(s))
            if len(unique_s) < 2:
                continue
            remaining = [idx for idx in unique_s if idx not in used_indices]
            if len(remaining) >= 2:
                effective_sets.append((gi, remaining))
                used_indices.update(remaining)

    # row_index -> (group_idx, sorted_set)。各インデックスは高々1つの set にのみ属する
    row_to_merge: Dict[int, Tuple[int, List[int]]] = {}
    for gi, set_list in effective_sets:
        for idx in set_list:
            row_to_merge[idx] = (gi, set_list)

    new_run_states: List[Dict[str, Any]] = []
    for i in range(len(run_states)):
        if i not in row_to_merge:
            new_run_states.append(run_states[i])
            continue
        gi, set_list = row_to_merge[i]
        if i != set_list[0]:
            continue
        rows_sub = [rows[j] for j in set_list]
        run_states_sub = [run_states[j] for j in set_list]
        order = sorted(range(len(rows_sub)), key=lambda j: _row_to_dt(rows_sub[j].get("出庫日時")) or datetime.min)
        rows_sub = [rows_sub[k] for k in order]
        run_states_sub = [run_states_sub[k] for k in order]
        merged_row, merged_rs = _merge_runs(rows_sub, run_states_sub, headers)
        if run_date_choices is not None and gi < len(run_date_choices):
            idx_choice = run_date_choices[gi]
            if isinstance(idx_choice, int) and 0 <= idx_choice < len(rows_sub):
                merged_row["運行日"] = rows_sub[idx_choice].get("運行日")
                merged_rs["merged_row"]["運行日"] = merged_row["運行日"]
                merged_rs["merged_header"]["運行日"] = merged_row["運行日"]
        new_run_states.append(merged_rs)

    new_rows: List[Dict[str, Any]] = []
    for rs in new_run_states:
        if rs.get("merged_row"):
            new_rows.append(dict(rs["merged_row"]))
        else:
            new_rows.append(row_from_run_state(rs))
    return new_run_states, new_rows


def apply_alcohol_to_run_states(
    run_states: List[Dict[str, Any]],
    alcohol_events: List[Any],
    margin_minutes: int = 120,
) -> None:
    """run_states の各 merged_header の出庫・帰庫をアルコール突合結果で上書きする（in place）。"""
    for rs in run_states:
        mh = rs.get("merged_header") or {}
        out_matched, in_matched = match_alcohol_for_run(
            alcohol_events,
            mh.get("乗務員ID"),
            mh.get("出庫日時"),
            mh.get("帰庫日時"),
            margin_minutes=margin_minutes,
        )
        mh["出庫日時"] = format_dt_for_excel(out_matched)
        mh["帰庫日時"] = format_dt_for_excel(in_matched)

        # ここが重要:
        # アルコール時刻で header を上書きしたら、
        # 以前の merged_row は時刻と計算値がズレるので必ず破棄する
        rs["merged_row"] = None

def rows_from_run_states(
    run_states: List[Dict[str, Any]],
    headers: List[str],
    preset_path: Path,
    device: str,
) -> List[Dict[str, Any]]:
    """run_states から Excel 行のリストを復元する（出庫・帰庫未取得判定などに使用）。"""
    preset = _load_preset(preset_path)
    ctx = {
        "timestamp": "",
        "company": "",
        "device_type": device,
        "report_id": "",
        "pdf_filename": "",
        "level": "",
        "category": "",
        "field_name": "",
        "value_candidates": "",
        "message": "",
    }

    def row_from_run_state(rs: Dict[str, Any]) -> Dict[str, Any]:
        merged_header = dict(rs.get("merged_header") or {})
        merged_details = rs.get("merged_details") or []

        cached_row = rs.get("merged_row")
        if cached_row:
            # キャッシュが header と一致している時だけ使う
            if (
                (cached_row.get("出庫日時") == merged_header.get("出庫日時"))
                and (cached_row.get("帰庫日時") == merged_header.get("帰庫日時"))
            ):
                return dict(cached_row)

        metrics = _compute_metrics(merged_header, merged_details, [], ctx, preset)
        row = {**merged_header, **metrics}

        # 統合行は _merge_runs 直後の運転時間を使う。再計算すると運行間ギャップを拾って膨張するため上書きしない
        prev_drive = merged_header.get("_merged_drive_min_initial")
        if prev_drive is not None:
            new_drive = metrics.get("運転時間")
            try:
                new_int = int(new_drive) if new_drive is not None else None
            except (TypeError, ValueError):
                new_int = None
            try:
                prev_int = int(prev_drive)
            except (TypeError, ValueError):
                prev_int = None
            _get_drive_compare_logger().info(
                "DRIVE_COMPARE run_id=%s initial=%s recomputed=%s (using initial)",
                merged_header.get("運行ID"),
                prev_int,
                new_int,
            )
            _apply_merged_drive_override(row, merged_header)

        if device in ("telecom", "mimamori") and row.get("運行ID"):
            rid = str(row["運行ID"])
            if rid.startswith("ID-"):
                row["運行ID"] = rid[3:]

        return row

    return [row_from_run_state(rs) for rs in run_states]


def _write_excel(headers: List[str], rows: List[Dict[str, Any]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "出力"

    ws.append(headers)

    header_to_colidx = {h: i + 1 for i, h in enumerate(headers)}

    # まず値を行単位で書く（h:mm列はここでは「分(int)」を一旦入れる）
    for r in rows:
        ws.append([r.get(h, None) for h in headers])

    # (0) ID列・安全点数を数値として変換（運行ID、乗務員ID、車両ID、営業所ID、安全点数）
    id_cols = {"運行ID", "乗務員ID", "車両ID", "営業所ID", "安全点数"}
    for h in id_cols:
        if h in header_to_colidx:
            col = header_to_colidx[h]
            for row_idx in range(2, ws.max_row + 1):
                c = ws.cell(row=row_idx, column=col)
                if c.value is not None and c.value != "":
                    try:
                        # 文字列の場合は数値に変換を試みる
                        if isinstance(c.value, str):
                            value_str = str(c.value).strip()
                            # 数値として解釈できる場合はintに変換
                            if value_str.isdigit():
                                c.value = int(value_str)
                            else:
                                # 数字部分を抽出して変換を試みる（例: "ID-39682" → 39682）
                                match = re.search(r'\d+', value_str)
                                if match:
                                    c.value = int(match.group())
                        elif isinstance(c.value, (int, float)):
                            c.value = int(c.value)
                    except (ValueError, TypeError):
                        # 変換できない場合はそのまま（文字列のまま）
                        pass

    # (0.5) 出庫メーター・帰庫メーター：小数点第1位まで、表示形式は km（全社共通）
    meter_cols = {"出庫メーター", "帰庫メーター"}
    for h in meter_cols:
        if h in header_to_colidx:
            col = header_to_colidx[h]
            for row_idx in range(2, ws.max_row + 1):
                c = ws.cell(row=row_idx, column=col)
                if c.value is not None and c.value != "":
                    try:
                        v = float(c.value)
                        c.value = round(v, 1)
                        c.number_format = '0.0 "km"'
                    except (ValueError, TypeError):
                        pass

    # (1) 分割開始/終了：datetime表示
    dt_cols = {"分割開始1", "分割終了1", "分割開始2", "分割終了2"}
    for h in dt_cols:
        if h in header_to_colidx:
            col = header_to_colidx[h]
            for row_idx in range(2, ws.max_row + 1):
                c = ws.cell(row=row_idx, column=col)
                if isinstance(c.value, datetime):
                    c.number_format = "yyyy/m/d h:mm"

    # (2) 「分(int)」で保持している列を Excelの time serial に変換し、表示を [h]:mm に統一
    #     ※セル表示だけを時間にし、内部計算はコード側で分に統一する、という方針
    #     ※休息時間・休息時間_昼・休息時間_夜は 0:00 の場合は転記しない（セルは空）
    rest_time_cols = {"休息時間", "休息時間_昼", "休息時間_夜"}
    for h in headers:
        if not _is_time_serial_col(h):
            continue
        col = header_to_colidx[h]
        for row_idx in range(2, ws.max_row + 1):
            c = ws.cell(row=row_idx, column=col)
            if isinstance(c.value, int):
                if h in rest_time_cols and c.value == 0:
                    c.value = None  # 0:00 の場合は転記しない
                else:
                    c.value = _minutes_to_excel_time_serial(c.value)
                    c.number_format = "[h]:mm"
            elif isinstance(c.value, float):
                # 万一floatで入っていても分として扱う
                mins = int(c.value)
                if h in rest_time_cols and mins == 0:
                    c.value = None  # 0:00 の場合は転記しない
                else:
                    c.value = _minutes_to_excel_time_serial(mins)
                    c.number_format = "[h]:mm"

    # データ全体のフォントを游ゴシックに統一
    font_yu = Font(name="游ゴシック")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = font_yu

    wb.save(out_path)


def _write_log(log_path: Path, entries: List[Dict[str, Any]]) -> None:
    with open(log_path, "w", encoding="utf-8-sig", newline="") as f:
        fieldnames = [
            "timestamp",
            "company",
            "device_type",
            "report_id",
            "pdf_filename",
            "level",
            "category",
            "field_name",
            "value_candidates",
            "message",
        ]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for e in entries:
            w.writerow({k: e.get(k, "") for k in fieldnames})


def _write_skipped(skipped_path: Path, skipped: List[Dict[str, Any]]) -> None:
    skipped_path.write_text(json.dumps(skipped, ensure_ascii=False, indent=2), encoding="utf-8")

def _merge_header_preferring_left(a: Dict[str, Any], b: Dict[str, Any]) -> Dict[str, Any]:
    """
    C案：矛盾したらエラー扱い（ログに残して空欄）に寄せるため、
    ここでは「同じキーで両方に値があり、かつ値が異なる」場合は None にする。
    """
    out = dict(a)
    for k, v in b.items():
        if k not in out or out[k] in (None, "", 0):
            out[k] = v
        else:
            # 両方に値があり、異なるなら矛盾 → None
            if v not in (None, "") and out[k] != v:
                out[k] = None
    return out


def _normalize_crew_id(uid: Any) -> str:
    """乗務員ID比較用（先頭0除去）。"""
    s = str(uid).strip() if uid is not None else ""
    if not s:
        return ""
    return s.lstrip("0") or "0"


GAP_MERGE_THRESHOLD_MINUTES = 3 * 60  # 帰庫→次出庫が3時間未満なら「1つにまとめますか？」の対象


def _detect_merge_groups(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    同一乗務員で、帰庫日時から次の出庫日時の間が3時間未満の運行を連鎖でグループ化する。
    各グループに運行日リスト（運行日リスト）を入れ、統合時にユーザーがどれを採用するか選べるようにする。
    """
    from collections import defaultdict
    by_crew: Dict[str, List[Tuple[int, str, Any]]] = defaultdict(list)  # 乗務員ID正規化 -> [(rowIndex, 運行ID, 運行日), ...]
    for i, r in enumerate(rows):
        cid = _normalize_crew_id(r.get("乗務員ID"))
        rid = str(r.get("運行ID") or "").strip()
        run_date = r.get("運行日")
        by_crew[cid].append((i, rid, run_date))
    out: List[Dict[str, Any]] = []
    for cid, list_with_dates in by_crew.items():
        if len(list_with_dates) < 2:
            continue
        # 出庫日時で昇順ソート
        def sort_key(item: Tuple[int, str, Any]) -> datetime:
            idx, _, _ = item
            t = _row_to_dt(rows[idx].get("出庫日時"))
            return t or datetime.min
        list_with_dates = sorted(list_with_dates, key=sort_key)
        # 連鎖: 隣同士で (次の出庫 - 前の帰庫) < 3h の塊に分割
        chains: List[List[Tuple[int, str, Any]]] = []
        cur_chain: List[Tuple[int, str, Any]] = [list_with_dates[0]]
        for k in range(1, len(list_with_dates)):
            prev_idx = list_with_dates[k - 1][0]
            curr_idx = list_with_dates[k][0]
            prev_in = _row_to_dt(rows[prev_idx].get("帰庫日時"))
            curr_out = _row_to_dt(rows[curr_idx].get("出庫日時"))
            gap_min = None
            if prev_in is not None and curr_out is not None and curr_out > prev_in:
                gap_min = int((curr_out - prev_in).total_seconds() // 60)
            if gap_min is not None and gap_min < GAP_MERGE_THRESHOLD_MINUTES:
                cur_chain.append(list_with_dates[k])
            else:
                chains.append(cur_chain)
                cur_chain = [list_with_dates[k]]
        if cur_chain:
            chains.append(cur_chain)
        for chain in chains:
            if len(chain) < 2:
                continue
            run_ids = [c[1] for c in chain]
            if len(set(run_ids)) < 2:
                continue
            row_indices = [c[0] for c in chain]
            run_dates = [c[2] for c in chain]
            first_row = rows[row_indices[0]]
            運行リスト = [
                {
                    "運行ID": rows[idx].get("運行ID"),
                    "出庫日時": rows[idx].get("出庫日時") or "",
                    "帰庫日時": rows[idx].get("帰庫日時") or "",
                    "運行日": rows[idx].get("運行日"),
                }
                for idx in row_indices
            ]
            out.append({
                "rowIndices": row_indices,
                "運行IDs": run_ids,
                "運行リスト": 運行リスト,
                "乗務員ID": first_row.get("乗務員ID"),
                "乗務員名": first_row.get("乗務員名"),
                "運行日リスト": run_dates,
            })
    return out


# 統合時に合算する列（N～AW のうち、分割・安全点数・AN～APを除く）
_MERGE_SUM_COLUMNS = [
    "拘束時間_分割前", "拘束時間_昼_分割前", "拘束時間_夜_分割前",
    "拘束時間_分割後", "拘束時間_昼_分割後", "拘束時間_夜_分割後",
    "労働時間_分割前", "労働時間_昼_分割前", "労働時間_夜_分割前",
    "労働時間_分割後", "労働時間_昼_分割後", "労働時間_夜_分割後",
    "運転時間", "待機時間", "荷積時間", "荷卸時間", "作業時間",
    "休憩時間_分割前", "休憩時間_昼_分割前", "休憩時間_夜_分割前",
    "休憩時間_分割後", "休憩時間_昼_分割後", "休憩時間_夜_分割後",
    "休息時間", "休息時間_昼", "休息時間_夜",
    "ランキング",
]
_MERGE_SUM_COLUMNS_SET = set(_MERGE_SUM_COLUMNS)
# 休憩に加算する列（運行間<3h と 分割から外れた区間）
_MERGE_KEIKAI_COLS = ("休憩時間_分割後", "休憩時間_昼_分割後", "休憩時間_夜_分割後")
# 休息に加算する列（運行間>=3h で分割から外れた場合）
_MERGE_KYUSOKU_COLS = ("休息時間", "休息時間_昼", "休息時間_夜")
# 分割列（上位2つを採用）
_MERGE_BUNKATSU_KEYS = ("分割開始1", "分割終了1", "分割1_作業時間_分", "分割開始2", "分割終了2", "分割2_作業時間_分")
# 同一運行としてまとめた行で Excel に「0」を転記しない列（Q～S, W～Y, AH～AJ, AN～AP）
_MERGE_BLANK_COLUMNS = (
    "拘束時間_分割後", "拘束時間_昼_分割後", "拘束時間_夜_分割後",
    "労働時間_分割後", "労働時間_昼_分割後", "労働時間_夜_分割後",
    "休憩時間_分割後", "休憩時間_昼_分割後", "休憩時間_夜_分割後",
    "休息時間採用", "運行間休息", "ランキング",
)
REST_THRESHOLD_MINUTES = 3 * 60  # 3時間以上 → 休息


def _merge_runs(rows: List[Dict[str, Any]], run_states: List[Dict[str, Any]], headers: List[str]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """
    同一乗務員の複数運行を1運行に統合する。
    rows, run_states は出庫日時で昇順に並んでいる前提。
    戻り値: (merged_row, merged_run_state)
    """
    def num_val(r: Dict[str, Any], k: str) -> int:
        v = r.get(k)
        if v is None or v == "":
            return 0
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return 0

    def float_val(r: Dict[str, Any], k: str) -> float:
        v = r.get(k)
        if v is None or v == "":
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    first = rows[0]
    merged_row: Dict[str, Any] = {}

    # 運行ID・運行日・乗務員は先頭採用
    merged_row["運行ID"] = first.get("運行ID")
    merged_row["運行日"] = first.get("運行日")
    merged_row["乗務員ID"] = first.get("乗務員ID")
    merged_row["乗務員名"] = first.get("乗務員名")
    merged_row["車両ID"] = first.get("車両ID")
    merged_row["車両番号"] = first.get("車両番号")
    merged_row["営業所ID"] = first.get("営業所ID")
    merged_row["所属営業所"] = first.get("所属営業所")

    # 出庫・帰庫は min / max
    out_dts = [_row_to_dt(r.get("出庫日時")) for r in rows]
    in_dts = [_row_to_dt(r.get("帰庫日時")) for r in rows]
    out_dts = [t for t in out_dts if t is not None]
    in_dts = [t for t in in_dts if t is not None]
    merged_row["出庫日時"] = format_dt_for_excel(min(out_dts)) if out_dts else first.get("出庫日時")
    merged_row["帰庫日時"] = format_dt_for_excel(max(in_dts)) if in_dts else first.get("帰庫日時")

    # メーター
    meter_out = [float_val(r, "出庫メーター") for r in rows]
    meter_in = [float_val(r, "帰庫メーター") for r in rows]
    merged_row["出庫メーター"] = min(meter_out) if meter_out else first.get("出庫メーター")
    merged_row["帰庫メーター"] = max(meter_in) if meter_in else first.get("帰庫メーター")

    # 総走行距離は合算（帰庫－出庫で再計算でもよいが、ここでは合算）
    merged_row["総走行距離"] = sum(float_val(r, "総走行距離") for r in rows) if rows else 0

    # 合算列
    for col in _MERGE_SUM_COLUMNS:
        merged_row[col] = sum(num_val(r, col) for r in rows)

    # 運行間の区間を列挙（出庫順に並んでいるので、隣接ペアの間が運行間）
    gap_segments: List[Dict[str, Any]] = []  # { start, end, dur_min, day_min, night_min }
    for idx in range(len(rows) - 1):
        r1, r2 = rows[idx], rows[idx + 1]
        t_end = _row_to_dt(r1.get("帰庫日時"))
        t_start = _row_to_dt(r2.get("出庫日時"))
        if t_end is None or t_start is None or t_start <= t_end:
            continue
        dur = int((t_start - t_end).total_seconds() // 60)
        day_min, night_min = _split_day_night_minutes(t_end, t_start)
        gap_segments.append({
            "start": t_end, "end": t_start, "dur": dur,
            "day_min": day_min, "night_min": night_min,
        })

    # 各運行の分割1・2を候補に、運行間(>=3h)も候補に
    candidates: List[Dict[str, Any]] = []  # { dur, day_min, night_min, start, end, is_gap }
    for r in rows:
        for i in (1, 2):
            start = r.get(f"分割開始{i}")
            end = r.get(f"分割終了{i}")
            dur = num_val(r, f"分割{i}_作業時間_分")
            if dur <= 0 and start is None and end is None:
                continue
            start_dt = _row_to_dt(start) if start is not None else None
            end_dt = _row_to_dt(end) if end is not None else None
            if start_dt and end_dt:
                day_min, night_min = _split_day_night_minutes(start_dt, end_dt)
            else:
                day_min, night_min = 0, 0
            candidates.append({
                "dur": dur, "day_min": day_min, "night_min": night_min,
                "start": start, "end": end, "start_dt": start_dt, "end_dt": end_dt,
                "is_gap": False,
            })
    # 運行間時間の扱い
    # - 3時間以上: 分割休息候補に入れる
    # - 3時間未満: 休憩時間_分割前 に直接加算する
    extra_gap_break_day = 0
    extra_gap_break_night = 0

    for g in gap_segments:
        if g["dur"] >= REST_THRESHOLD_MINUTES:
            candidates.append({
                "dur": g["dur"],
                "day_min": g["day_min"],
                "night_min": g["night_min"],
                "start": g["start"],
                "end": g["end"],
                "start_dt": g["start"],
                "end_dt": g["end"],
                "is_gap": True,
            })
        else:
            extra_gap_break_day += g["day_min"]
            extra_gap_break_night += g["night_min"]

    # 上位2つを分割①・②に
    candidates.sort(key=lambda x: x["dur"], reverse=True)
    top2 = candidates[:2]
    for i, c in enumerate(top2, start=1):
        st = c.get("start_dt") or c.get("start")
        en = c.get("end_dt") or c.get("end")
        if hasattr(st, "strftime"):
            st = st.strftime("%Y/%m/%d %H:%M") if st else None
        if hasattr(en, "strftime"):
            en = en.strftime("%Y/%m/%d %H:%M") if en else None
        merged_row[f"分割開始{i}"] = st
        merged_row[f"分割終了{i}"] = en
        merged_row[f"分割{i}_作業時間_分"] = c["dur"]

    # 休息（AK~AM）は分割①②の2つだけとする
    rest_day = sum(c["day_min"] for c in top2)
    rest_night = sum(c["night_min"] for c in top2)
    merged_row["休息時間"] = rest_day + rest_night
    merged_row["休息時間_昼"] = rest_day
    merged_row["休息時間_夜"] = rest_night

    # 分割から外れた区間 + 3時間未満の運行間時間 は休憩時間_分割前に加算
    fallen = candidates[2:]
    extra_keikai_pre_day, extra_keikai_pre_night = 0, 0
    for c in fallen:
        extra_keikai_pre_day += c["day_min"]
        extra_keikai_pre_night += c["night_min"]

    extra_keikai_pre_day += extra_gap_break_day
    extra_keikai_pre_night += extra_gap_break_night

    merged_row["休憩時間_分割前"] = (
        num_val(merged_row, "休憩時間_分割前")
        + extra_keikai_pre_day
        + extra_keikai_pre_night
    )
    merged_row["休憩時間_昼_分割前"] = (
        num_val(merged_row, "休憩時間_昼_分割前")
        + extra_keikai_pre_day
    )
    merged_row["休憩時間_夜_分割前"] = (
        num_val(merged_row, "休憩時間_夜_分割前")
        + extra_keikai_pre_night
    )

    # 安全点数: 平均して四捨五入（1件でもあれば採用、なければ先頭行から転記）
    safe_scores = [float_val(r, "安全点数") for r in rows if r.get("安全点数") not in (None, "")]
    if safe_scores:
        merged_row["安全点数"] = round(sum(safe_scores) / len(safe_scores))
    else:
        # 数値化できなかった場合も、いずれかの行に値があれば転記する（みまもり等で合算時に欠落しないように）
        merged_row["安全点数"] = next(
            (r.get("安全点数") for r in rows if r.get("安全点数") not in (None, "")),
            first.get("安全点数"),
        )

    # 同一運行まとめ行で転記しない列は None にし、Excel では空白になる
    for col in _MERGE_BLANK_COLUMNS:
        merged_row[col] = None

    # merged_run_state: 統合行を再計算に使うため merged_header + merged_details（安全点数も持たせておく）
    first_rs = run_states[0]
    merged_header = {
        "運行ID": merged_row["運行ID"],
        "運行日": merged_row["運行日"],
        "乗務員ID": merged_row["乗務員ID"],
        "乗務員名": merged_row["乗務員名"],
        "車両ID": merged_row["車両ID"],
        "車両番号": merged_row["車両番号"],
        "営業所ID": merged_row["営業所ID"],
        "所属営業所": merged_row["所属営業所"],
        "出庫メーター": merged_row["出庫メーター"],
        "帰庫メーター": merged_row["帰庫メーター"],
        "出庫日時": merged_row["出庫日時"],
        "帰庫日時": merged_row["帰庫日時"],
        "安全点数": merged_row.get("安全点数"),
        # _merge_runs 直後の運転時間（分）。後続で rows_from_run_states() による再計算が入ったか比較するためのログ用。
        "_merged_drive_min_initial": merged_row.get("運転時間"),
    }
    # 3時間以上紐づけのドロップダウン表示用：統合元のデジタコ出庫・帰庫の min/max を退避
    out_digitaco_dts = []
    in_digitaco_dts = []
    for rs in run_states:
        mh = rs.get("merged_header") or {}
        o = mh.get("_digitaco_出庫日時") or mh.get("出庫日時")
        i = mh.get("_digitaco_帰庫日時") or mh.get("帰庫日時")
        if o:
            t = _row_to_dt(o)
            if t is not None:
                out_digitaco_dts.append(t)
        if i:
            t = _row_to_dt(i)
            if t is not None:
                in_digitaco_dts.append(t)
    if out_digitaco_dts:
        merged_header["_digitaco_出庫日時"] = format_dt_for_excel(min(out_digitaco_dts))
    if in_digitaco_dts:
        merged_header["_digitaco_帰庫日時"] = format_dt_for_excel(max(in_digitaco_dts))

    all_details: List[Dict[str, Any]] = []
    seq_no = 0
    for run_order, rs in enumerate(run_states):
        details = rs.get("merged_details") or []
        for detail_order, d in enumerate(details):
            nd = dict(d)
            nd["_merge_run_order"] = run_order
            nd["_merge_detail_order"] = detail_order
            nd["_merge_seq"] = seq_no
            seq_no += 1
            all_details.append(nd)

    merged_run_state = {
        "report_id": first_rs.get("report_id"),
        "merged_header": merged_header,
        "merged_details": all_details,
        "merged_row": merged_row,
    }
    return merged_row, merged_run_state


def _merge_detail_rows(details_list: List[List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    """
    明細は単純に結合し、(item, task, arrival, depart) で重複排除。
    itemはPDF分割だと重複したり飛ぶので、ソートは item->arrival->depart で安定化。
    """
    merged: List[Dict[str, Any]] = []
    seen = set()
    for details in details_list:
        for r in details:
            key = (r.get("item"), r.get("task"), r.get("arrival"), r.get("depart"))
            if key in seen:
                continue
            seen.add(key)
            merged.append(r)

    def _key(r: Dict[str, Any]):
        item = r.get("item")
        item = item if isinstance(item, int) else 10**9
        return (item, r.get("arrival") or "99:99", r.get("depart") or "99:99")

    return sorted(merged, key=_key)
# =========================
# Entry
# =========================

def run_pipeline(
    *,
    company: str,
    device: str,            # "mimamori" | "telecom"
    preset_path: Path,      # companies/<company>/<device>.json
    pdf_paths: List[Path],
    job_output_dir: Path,
    job_input_dir: Optional[Path] = None,  # アルコール突合用（taimen / alcohol サブディレクトリ）
) -> PipelineResult:
    from storage.paths import EXCEL_HEADERS_JSON_PATH
    _headers_path = EXCEL_HEADERS_JSON_PATH if EXCEL_HEADERS_JSON_PATH is not None else (Path(__file__).parent / "excel_headers.json")
    headers: List[str] = json.loads(_headers_path.read_text(encoding="utf-8"))
    preset = _load_preset(preset_path)

    job_output_dir.mkdir(parents=True, exist_ok=True)

    excel_path = job_output_dir / "output.xlsx"
    log_path = job_output_dir / "log.csv"
    skipped_path = job_output_dir / "skipped.json"

    rows: List[Dict[str, Any]] = []
    logs: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []

    error_count = 0
    warn_count = 0

    # ---- 1) まず各PDFを読み、運行IDごとにグルーピング ----
    groups: Dict[str, List[Dict[str, Any]]] = {}  # report_id -> list of parts
    unknown_idx = 0

    for p in pdf_paths:
        ctx_base = {
            "timestamp": _now_iso(),
            "company": company,
            "device_type": device,
            "report_id": "",
            "pdf_filename": p.name,
            "level": "",
            "category": "",
            "field_name": "",
            "value_candidates": "",
            "message": "",
        }

        try:
            raw, _ = _read_pdf_text(p)
            he = preset.get("header_extract") or {}
            report_id_regex = he.get("report_id_regex") or r"ID-\d+"
            run_blocks = _split_raw_by_runs(raw, report_id_regex)

            for run_block in run_blocks:
                cleaned_block = _clean_for_regex(run_block)
                header = _extract_header_fields(cleaned_block, device, preset)
                report_id = header.get("運行ID") or ""

                # 運行IDが取れないブロックは「unknown_xxx」として単独処理（スキップはしない）
                if not report_id:
                    unknown_idx += 1
                    report_id = f"unknown_{unknown_idx}"
                    warn_count += 1
                    logs.append({**ctx_base, "level": "WARN", "category": "HEADER_WEAK", "field_name": "運行ID", "message": "運行IDが取れないため単独処理"})

                detail_rows = _extract_detail_rows(run_block)

                part = {
                    "pdf": p.name,
                    "raw": run_block,
                    "cleaned": cleaned_block,
                    "header": header,
                    "details": detail_rows,
                }
                groups.setdefault(report_id, []).append(part)

        except Exception as e:
            error_count += 1
            skipped.append({"pdf": p.name, "reason": str(e)})
            logs.append({
                **ctx_base,
                "level": "ERROR",
                "category": "PIPELINE_EXCEPTION",
                "field_name": "",
                "value_candidates": "",
                "message": f"例外によりスキップ: {e}",
            })

    # アルコール突合用に事前に統合イベントを読み込む（job_input_dir がある場合のみ）
    alcohol_events: Optional[List[Any]] = None
    if job_input_dir is not None:
        taimen_dir = job_input_dir / "taimen"
        alcohol_dir = job_input_dir / "alcohol"
        alcohol_events = integrate_alcohol(taimen_dir, alcohol_dir)

    run_states: List[Dict[str, Any]] = []  # 手入力完了時に再計算する用（alcohol 使用時のみ蓄積）

    # ---- 2) グループごとに結合して 1運行=1行 を作る ----
    for report_id, parts in groups.items():
        ctx = {
            "timestamp": _now_iso(),
            "company": company,
            "device_type": device,
            "report_id": report_id if not report_id.startswith("unknown_") else "",
            "pdf_filename": ",".join([p["pdf"] for p in parts]),
            "level": "",
            "category": "",
            "field_name": "",
            "value_candidates": "",
            "message": "",
        }

        try:
            # header 結合（矛盾は None）
            merged_header: Dict[str, Any] = {}
            for part in parts:
                merged_header = _merge_header_preferring_left(merged_header, part["header"])

            # C案：矛盾して None になった主要項目がある場合はログに残す（空欄のまま続行）
            for k in ("運行日", "乗務員ID", "車両ID", "出庫日時", "帰庫日時"):
                if k in merged_header and merged_header[k] is None:
                    error_count += 1
                    logs.append({**ctx, "level": "ERROR", "category": "HEADER_CONFLICT", "field_name": k, "message": "同一運行ID内で値が矛盾。空欄扱い"})

            # ②をアルコール突合より前にするため、ここではアルコール突合しない（後で merge がなければ一括適用する）
            # details 結合
            merged_details = _merge_detail_rows([p["details"] for p in parts])

            # 手入力用に run_state を蓄積（JSON 化できる形で）。デジタコの出庫・帰庫のみ入れた状態で保存。
            # アルコール突合で上書きされても、3h以上紐づけ画面でデジタコの出庫・帰庫を表示するため退避用に保存
            def _to_serializable(v: Any) -> Any:
                if isinstance(v, datetime):
                    return v.strftime("%Y/%m/%d %H:%M")
                return v
            mh_serialized = {k: _to_serializable(v) for k, v in merged_header.items()}
            mh_serialized["_digitaco_出庫日時"] = mh_serialized.get("出庫日時")
            mh_serialized["_digitaco_帰庫日時"] = mh_serialized.get("帰庫日時")
            run_states.append({
                "report_id": report_id,
                "merged_header": mh_serialized,
                "merged_details": [{k: _to_serializable(v) for k, v in d.items()} for d in merged_details],
            })

            # metrics 計算（出庫・帰庫はデジタコのまま。欠損なら拘束時間・労働時間等は取れず空になる）
            metrics = _compute_metrics(merged_header, merged_details, logs, ctx, preset)

            row: Dict[str, Any] = {}
            row.update(merged_header)
            row.update(metrics)

            # テレコム・みまもり機種の場合、運行IDから「ID-」プレフィックスを削除
            if device in ("telecom", "mimamori") and "運行ID" in row and row["運行ID"]:
                run_id = str(row["運行ID"])
                if run_id.startswith("ID-"):
                    row["運行ID"] = run_id[3:]  # "ID-" の3文字を削除

            # ここで運行IDを確実に入れる（unknown は空欄でもOKなら消す）
            if not report_id.startswith("unknown_"):
                # テレコム・みまもり機種の場合、「ID-」プレフィックスを削除
                final_report_id = report_id
                if device in ("telecom", "mimamori") and report_id.startswith("ID-"):
                    final_report_id = report_id[3:]  # "ID-" の3文字を削除
                row["運行ID"] = final_report_id

            rows.append(row)

        except Exception as e:
            error_count += 1
            pdf_names = ",".join([p["pdf"] for p in parts])
            skipped.append({"pdf": pdf_names, "reason": str(e)})
            logs.append({
                **ctx,
                "level": "ERROR",
                "category": "PIPELINE_EXCEPTION",
                "field_name": "",
                "value_candidates": "",
                "message": f"例外によりスキップ: {e}",
            })

    _write_log(log_path, logs)
    _write_skipped(skipped_path, skipped)

    # ②をアルコール突合より前に: 3h未満グループがあればここで質問に回し、アルコールはまだかけない
    if run_states and rows:
        merge_groups = _detect_merge_groups(rows)
        if merge_groups and alcohol_events is not None:
            return PipelineResult(
                excel_path, log_path, skipped_path, error_count, warn_count,
                merge_decision_required=True,
                merge_groups=merge_groups,
                run_states=run_states,
                headers=headers,
            )

    # ここに来た場合は merge で return していない。アルコール突合を実行して rows を更新
    if alcohol_events is not None and run_states:
        apply_alcohol_to_run_states(run_states, alcohol_events, margin_minutes=120)
        rows[:] = rows_from_run_states(run_states, headers, preset_path, device)

    # アルコール使用かつ出庫・帰庫が取れていない行があれば手入力フローへ
    if alcohol_events is not None and run_states:
        def _run_date_for_row(r: Dict[str, Any]) -> Any:
            v = r.get("運行日")
            if v is not None and str(v).strip():
                return v
            dt = _row_to_dt(r.get("出庫日時") or r.get("帰庫日時"))
            return dt.strftime("%Y/%m/%d") if dt else None
        missing = [
            {"rowIndex": i, "運行ID": r.get("運行ID"), "乗務員ID": r.get("乗務員ID"), "乗務員名": r.get("乗務員名"), "運行日": _run_date_for_row(r), "出庫日時": r.get("出庫日時") or "", "帰庫日時": r.get("帰庫日時") or ""}
            for i, r in enumerate(rows)
            if not r.get("出庫日時") or not r.get("帰庫日時")
        ]
        if missing:
            alc_runs = get_alcohol_runs_by_crew(alcohol_events)
            return PipelineResult(
                excel_path, log_path, skipped_path, error_count, warn_count,
                manual_input_required=True,
                run_states=run_states,
                pending_rows=missing,
                headers=headers,
                alcohol_runs_by_crew=alc_runs,
            )

    _write_excel(headers, rows, excel_path)
    return PipelineResult(excel_path, log_path, skipped_path, error_count, warn_count)
